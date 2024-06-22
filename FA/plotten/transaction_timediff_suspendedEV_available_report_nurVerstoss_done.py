import pandas as pd
import os
from db_connection import get_db_connection
from datetime import timedelta
from mail_connection import send_email
from fpdf import FPDF
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.utils import get_column_letter

# Notwendige Bibliotheken: pyarrow; fastparquet


#Inhalt:
# 1. Zusammenfassung des Codes:

# 2. Datenbankverbindung wird hergestellt.
# 3. Für jede Transaktion innerhalb eines bestimmten Bereichs wird eine SQL-Abfrage ausgeführt.
# 4. Ergebnisse der Abfrage werden in einem DataFrame gespeichert.
# 5. Zeitdifferenzen zwischen bestimmten Statusmeldungen werden berechnet.
# 6. Bei einem Verstoß (Zeitdifferenz > 30 Minuten) werden die Details der Transaktion gesammelt.
# 7. Die Ergebnisse werden in eine Textdatei und eine geschützte Excel-Datei geschrieben.
# 8. Ein E-Mail-Bericht kann an einen Administrator gesendet werden.

# Verstoßüberprüfung, Schwellwert = 30 min 
def ueberpruefung(difference_minutes):
    if difference_minutes > 30: 
        return "Verstoß!"
    else:
        return "Kein Verstoß!"
    
# Datenbankverbindung aufbauen
db_conn = get_db_connection()

# Ergebnis-DataFrame initialisieren
all_results = []

transaction_start = 2850
transaction_end = 2870

# Für jede Transaktion von ... bis ... die Abfrage ausführen
for transaction_id in range(transaction_start, transaction_end):

    # SQL-Abfrage für die aktuelle Transaktion ausführen
    sql_query = f"""
    SELECT 
        t.transaction_pk AS Transaktionsnummer, 
        u.first_name AS Vorname, 
        u.last_name AS Nachname, 
        u.e_mail AS e_mail, 
        cs.status_timestamp, 
        cs.`status`, 
        cs.vendor_id
    FROM 
        transaction t 
        JOIN connector_meter_value c USING(transaction_pk) 
        JOIN connector_status cs ON cs.connector_pk = c.connector_pk  
        JOIN ocpp_tag AS o ON t.id_tag = o.id_tag
        JOIN user AS u ON o.ocpp_tag_pk = u.ocpp_tag_pk
    WHERE 
        transaction_pk = {transaction_id}
        AND cs.vendor_id = 'EVTEC'   
        AND c.measurand = 'Power.Active.Import'
        AND cs.`status` IN ('SuspendedEV' , 'Available')
        AND cs.status_timestamp >= t.start_timestamp 
        AND cs.status_timestamp <= DATE_ADD(t.start_timestamp, INTERVAL 5 HOUR)
    ORDER BY 
        value_timestamp ASC;
    """
    
    # SQL-Abfrage ausführen und Ergebnis in einen DataFrame laden
    df = pd.read_sql_query(sql_query, db_conn)

    if df.empty:
        continue  # Überspringen, wenn keine Daten für die Transaktion vorhanden sind

    # Konvertieren der status_timestamp-Spalte in datetime
    df['status_timestamp'] = pd.to_datetime(df['status_timestamp'])

    # Berechnen der Differenz
    try:
        suspended_time = df[df['status'] == 'SuspendedEV']['status_timestamp'].iloc[0]
        available_time = df[df['status'] == 'Available']['status_timestamp'][df['status_timestamp'] > suspended_time].iloc[0]
        difference = available_time - suspended_time

        # Unterschied in Minuten
        difference_minutes = round(difference.total_seconds() / 60, 1)

        # Unterschied im hh:mm:ss Format
        difference_str = str(difference)

        # Überprüfen, ob die Differenz größer als 30 Minuten ist
        if difference_minutes > 30:
            print(f"Verstoß bei Transaktion {transaction_id}!")

            # Ergebnisse speichern, nur wenn Verstoß vorliegt
            all_results.append({
                'Transaktions-ID': transaction_id,
                'Vorname': df['Vorname'].iloc[0],
                'Nachname': df['Nachname'].iloc[0],
                'E-Mail': df['e_mail'].iloc[0],
                'Zeit SuspendedEV': suspended_time,
                'Zeit Available': available_time,
                'Differenz': difference_str,
                'Differenz in Minuten': difference_minutes,
                'Verstoß vorliegend?': ueberpruefung(difference_minutes)
            })

    except IndexError:
        # Falls keine passenden Zeitstempel vorhanden sind, überspringen
        continue
    
# Ergebnisse in einen DataFrame umwandeln
results_df = pd.DataFrame(all_results)

# Sicherstellen, dass der Ordner 'output_files' existiert
output_folder = 'output_files'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Erstellen des Textinhalts
text_content = "Transaktionsbericht\n"
text_content += "=" * 20 + "\n\n"

for index, row in results_df.iterrows():
    text_content += (f"Transaktions-ID: {row['Transaktions-ID']}\n"
                     f"Vorname: {row['Vorname']}\n"
                     f"Nachname: {row['Nachname']}\n"
                     f"E-Mail: {row['E-Mail']}\n"
                     f"Zeit SuspendedEV: {row['Zeit SuspendedEV']}\n"
                     f"Zeit Available: {row['Zeit Available']}\n"
                     f"Differenz: {row['Differenz']}\n"
                     f"Differenz in Minuten: {row['Differenz in Minuten']}\n"
                     f"Verstoß vorliegend?: {row['Verstoß vorliegend?']}\n")
    text_content += "-" * 20 + "\n"

# Relativer Pfad zur Textdatei
# output_file_path = os.path.join(output_folder, 'Transaktionsbericht.txt')

# Textinhalt in die Datei schreiben
# with open(output_file_path, 'w') as file:
#     file.write(text_content)

# Speichern der Ergebnisse in eine Excel-Datei, nur wenn Verstöße vorliegen
vendor_id = 'EVTEC'  # Hersteller
if not results_df.empty:
    excel_file_path = f'C:/Users/Tamer/Desktop/project_x/project_x/FA/output_files/timediff_{vendor_id}_SuspendedEV_bis_Available.xlsx'
    results_df.to_excel(excel_file_path, index=False)
    
    # Excel-Datei schützen
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active
    
    # Spaltenbreite auf 20 Exceleinheiten setzen
    for col in sheet.columns:
        max_length = 20 
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length
    
    sheet.protection.sheet = True
    sheet.protection.password = 'your_password'  # Passwort setzen
    workbook.security = WorkbookProtection(workbookPassword='your_password', lockStructure=True, lockWindows=True)
    workbook.save(excel_file_path)
    
    print("Datei erfolgreich gespeichert und geschützt.")

# __________________________________________________________________________________#
# Mail wird an den Admin gesendet
# send_email(
#     subject="Report E-Mail",
#     body="Report der Transaktion von 2850 bis 2900",
#     # to_email="abdelrahmanta00@gmail.com",
#     to_email="frank.brosi@fkfs.de",
#     from_email="abdelrahmanta00@gmail.com",
#     smtp_server="smtp.gmail.com",
#     smtp_port=587,
#     login="abdelrahmanta00@gmail.com",
#     password="qogz mjeo mrfn iwky",  # generierte App-Passwort
#     attachment_path="C:/Users/Tamer/Desktop/project_x/project_x/FA/output_files/timediff_EVTEC_SuspendedEV_bis_Available.xlsx"
# )
# print(f"E-Mail wurde erfolgreich verschickt an {abdelrahmanta00@gmail.com}")
