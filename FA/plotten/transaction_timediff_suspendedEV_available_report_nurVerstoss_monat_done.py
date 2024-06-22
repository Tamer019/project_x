import pandas as pd
import os
from db_connection import get_db_connection
from datetime import datetime, timedelta
from mail_connection import send_email
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.utils import get_column_letter
import time


# Notwendige Bibliotheken: pyarrow; fastparquet

# Verstoßüberprüfung, Schwellwert = 30 min 
def ueberpruefung(difference_minutes):
    if difference_minutes > 30: 
        return "Verstoß!"
    else:
        return "Kein Verstoß!"

# Einzigartigen Dateinamen erstellen
def create_unique_filename(directory, base_filename, extension):
    counter = 1
    filename = f"{base_filename}.{extension}"
    while os.path.exists(os.path.join(directory, filename)):
        filename = f"{base_filename}_{counter}.{extension}"
        counter += 1
    return os.path.join(directory, filename)

# Funktion zum Hervorheben von Text
def highlight(text, color_code="93"):
    return f"\033[{color_code}m{text}\033[0m"

# Umschaltmöglichkeit für Datumseingabe
print(highlight("Wählen Sie den Modus: (1) Aktuelles Datum (2) Manuelles Datum"))
modus = input(highlight("Modus: "))

if modus == '1':
    # Aktuelles Datum und erster/letzter Tag des Monats
    today = datetime.today()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
#test
elif modus == '2': 
    # Manuelles Jahr und Monat abfragen
    jahr = int(input(highlight("Geben Sie das Jahr ein (z.B. 2023): ")))
    monat = int(input(highlight("Geben Sie den Monat ein (1-12): ")))
    first_day_of_month = datetime(jahr, monat, 17)
    last_day_of_month = datetime(jahr, monat, 18)   #1 Tag mehr als erwümscht
# elif modus == '2':
#     # Manuelles Jahr und Monat abfragen
#     jahr = int(input(highlight("Geben Sie das Jahr ein (z.B. 2023): ")))
#     monat = int(input(highlight("Geben Sie den Monat ein (1-12): ")))
#     first_day_of_month = datetime(jahr, monat, 1)
#     last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
else:
    print("Ungültige Eingabe. Das Skript wird beendet.")
    exit()

# Startzeit des Skripts
start_time = time.time()

# Datenbankverbindung aufbauen
db_conn = get_db_connection()

# Ergebnis-DataFrame initialisieren
all_results = []

# # Aktuelles Datum und erster/letzter Tag des Monats
# # today = datetime.today()
# # first_day_of_month = today.replace(day=1)
# # last_day_of_month = (first_day_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)

# # Manuell gesetztes Datum für April 2023
# first_day_of_month = datetime(2023, 4, 15)
# last_day_of_month = datetime(2023, 4, 20)

# # Umschaltmöglicheit mit 1 2 oder so 

# SQL-Abfrage für Transaktionen des aktuellen Monats
sql_query = f"""
SELECT 
    t.transaction_pk AS Transaktionsnummer, 
    u.first_name AS Vorname, 
    u.last_name AS Nachname, 
    u.e_mail AS e_mail, 
    cs.status_timestamp, 
    cs.`status`, 
    cs.vendor_id
FROM 
    transaction t 
    JOIN connector_meter_value c USING(transaction_pk) 
    JOIN connector_status cs ON cs.connector_pk = c.connector_pk  
    JOIN ocpp_tag AS o ON t.id_tag = o.id_tag
    JOIN user AS u ON o.ocpp_tag_pk = u.ocpp_tag_pk
WHERE 
    t.start_timestamp >= '{first_day_of_month.strftime('%Y-%m-%d')}'
    AND t.start_timestamp <= '{last_day_of_month.strftime('%Y-%m-%d')}'
    AND cs.vendor_id = 'EVTEC'   
    AND c.measurand = 'Power.Active.Import'
    AND cs.`status` IN ('SuspendedEV' , 'Available')
    AND cs.status_timestamp >= t.start_timestamp 
    AND cs.status_timestamp <= DATE_ADD(t.start_timestamp, INTERVAL 5 HOUR)
ORDER BY 
    cs.status_timestamp ASC;
"""

# SQL-Abfrage ausführen und Ergebnis in einen DataFrame laden
df = pd.read_sql_query(sql_query, db_conn)

if not df.empty:
    # Konvertieren der status_timestamp-Spalte in datetime
    df['status_timestamp'] = pd.to_datetime(df['status_timestamp'])

    grouped = df.groupby('Transaktionsnummer')

    for transaction_id, group in grouped:
        try:
            suspended_time = group[group['status'] == 'SuspendedEV']['status_timestamp'].iloc[0]
            available_time = group[group['status'] == 'Available']['status_timestamp'][group['status_timestamp'] > suspended_time].iloc[0]
            difference = available_time - suspended_time

            # Unterschied in Minuten
            difference_minutes = round(difference.total_seconds() / 60, 1)

            # Unterschied im hh:mm:ss Format
            difference_str = str(difference)

            # Überprüfen, ob die Differenz größer als 30 Minuten ist
            if difference_minutes > 30:
                print(f"Verstoß bei Transaktion {transaction_id}!")

                # Ergebnisse speichern, nur wenn Verstoß vorliegt
                all_results.append({
                    'Transaktions-ID': transaction_id,
                    'Vorname': group['Vorname'].iloc[0],
                    'Nachname': group['Nachname'].iloc[0],
                    'E-Mail': group['e_mail'].iloc[0],
                    'Zeit SuspendedEV': suspended_time,
                    'Zeit Available': available_time,
                    'Differenz': difference_str,
                    'Differenz in Minuten': difference_minutes,
                    'Verstoß vorliegend?': ueberpruefung(difference_minutes)
                })

        except IndexError:
            # Falls keine passenden Zeitstempel vorhanden sind, überspringen
            continue

# Ergebnisse in einen DataFrame umwandeln
results_df = pd.DataFrame(all_results)

# Sicherstellen, dass der Ordner 'output_files' existiert
output_folder = 'output_files'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# Speichern der Ergebnisse in eine Excel-Datei, nur wenn Verstöße vorliegen
vendor_id = 'EVTEC'  # Hersteller
if not results_df.empty:
    excel_file_path = create_unique_filename(output_folder, f'timediff_Ladeende_Ausstecken', 'xlsx')
    results_df.to_excel(excel_file_path, index=False)

    # Excel-Datei schützen
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Spaltenbreite auf 20 Exceleinheiten setzen
    for col in sheet.columns:
        max_length = 20 
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length

    sheet.protection.sheet = True
    sheet.protection.password = 'your_password'  # Passwort setzen
    workbook.security = WorkbookProtection(workbookPassword='your_password', lockStructure=True, lockWindows=True)
    workbook.save(excel_file_path)

print("Datei erfolgreich gespeichert und geschützt.")

# Endzeit des Skripts
end_time = time.time()

# Zeitdifferenz berechnen
elapsed_time = end_time - start_time
print(f"Das Skript hat {elapsed_time:.2f} Sekunden benötigt.")

# __________________________________________________________________________________#
# Mail wird an den Admin gesendet
# send_email(
#     subject="Report E-Mail",
#     body="Bericht der Transaktionen für den angegebenen Zeitraum",
#     to_email="frank.brosi@fkfs.de",
#     from_email="abdelrahmanta00@gmail.com",
#     smtp_server="smtp.gmail.com",
#     smtp_port=587,
#     login="abdelrahmanta00@gmail.com",
#     password="qogz mjeo mrfn iwky",  # generierte App-Passwort
#     attachment_path=excel_file_path
# )
# print(f"E-Mail wurde erfolgreich verschickt an frank.brosi@fkfs.de")
